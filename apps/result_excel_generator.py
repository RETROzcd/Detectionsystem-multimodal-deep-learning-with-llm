import openpyxl
from openpyxl.drawing.image import Image
from PIL import Image as PILImage
import io
import base64
import logging
import os

class ExcelGenerator:
    def __init__(self):
        self.logger = logging.getLogger(__name__)
    
    def generate_review_excel(self, rule_check_response, manual_check_groups, output_path="审核结果.xlsx"):
        try:
            # 创建新的Excel工作簿
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "审核结果"
            
            # 设置列标题
            headers = ["章节", "标题", "方法", "要求", "交互中的“前置条件”匹配", "审核内容和规则判断", "豁免条款", "AI输出规则", "AI检查结果", "人工修正", "AI判断依据的参考图"]
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # 处理审核输出数据
            row = 2
            for i in range(0, len(rule_check_response.get_check_results())):  # 每组3个输出
                # 提取数据
                # if review_outputs[i] is None or len(review_outputs[i]) == 0:
                #     break
                check_result = rule_check_response.get_check_results()[i]
                chapter = check_result['rule'].chapter #章节
                title = check_result['rule'].title #标题
                method = check_result['rule'].method #方法
                requirements = check_result['rule'].requirements #要求
                preconditions = check_result['rule'].preconditions #前置条件
                exemption_clauses = check_result['rule'].exemption_clauses #豁免条款
                audit_content = check_result['rule'].audit_content #审核内容
                pics = check_result['check_result'].pics
                display_result = "NA"
                if check_result['check_result'].pass_status is None:
                    display_result = "NA"
                elif check_result['check_result'].pass_status:
                    display_result = "Pass"
                else:
                    display_result = "Fail"
                llm_response = f"{display_result}，{check_result['check_result'].llm_response}" if check_result['check_result'].necessity_state else "NA"
                llm_prompt = check_result['rule'].llm_prompt
                manual_correct_conclusion = manual_check_groups[i*3+0]
                manual_error_reason = manual_check_groups[i*3+1]
                manual_is_error = manual_check_groups[i*3+2]
                
                # 写入数据
                ws.cell(row=row, column=1, value=chapter) #章节 
                ws.cell(row=row, column=2, value=title) #标题
                ws.cell(row=row, column=3, value=method) #方法
                ws.cell(row=row, column=4, value=requirements) #要求
                ws.cell(row=row, column=5, value=preconditions) #交互中的“前置条件”匹配
                ws.cell(row=row, column=6, value=audit_content) #审核内容和规则判断
                ws.cell(row=row, column=7, value=exemption_clauses) #豁免条款
                ws.cell(row=row, column=8, value=llm_prompt) #AI输出规则
                ws.cell(row=row, column=9, value=llm_response) #AI检查结果
                # ws.cell(row=row, column=10, value="") #AI判断依据的参考图
                ws.cell(row=row, column=10, value=f"{manual_correct_conclusion}，{manual_error_reason}，{manual_is_error}") #人工修正
                for j, pic in enumerate(pics):
                    self._add_image_to_excel(ws, pic, 11+j, row) #AI判断依据的参考图
                
                row += 1
            
            # 保存Excel文件
            wb.save(output_path)
            self.logger.info(f"Excel文件已生成: {output_path}")
            
            return output_path
            
        except Exception as e:
            self.logger.error(f"生成Excel文件时出错: {str(e)}")
            raise
    
    def _add_image_to_excel(self, worksheet, img_data, col, row):
        try:
            # 获取图片数据
            if isinstance(img_data, str) and img_data.startswith('data:image'):
                # 处理base64图片
                img_data = img_data.split(',')[1]
                img_bytes = base64.b64decode(img_data)
            else:
                # 处理文件路径
                with open(img_data, 'rb') as f:
                    img_bytes = f.read()
            
            # 创建PIL图片并调整大小
            pil_img = PILImage.open(io.BytesIO(img_bytes))
            pil_img.thumbnail((200, 200))  # 调整图片大小
            
            # 保存调整后的图片到字节流
            img_buffer = io.BytesIO()
            pil_img.save(img_buffer, format='PNG')
            img_buffer.seek(0)
            
            # 插入图片到Excel
            img = Image(img_buffer)
            img.width = 150
            img.height = 150
            
            # 计算图片位置
            cell_address = f'{openpyxl.utils.get_column_letter(col)}{row}'
            worksheet.add_image(img, cell_address)
            
        except Exception as e:
            self.logger.warning(f"处理图片时出错: {str(e)}")

if __name__ == "__main__":
    excel_generator = ExcelGenerator()
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image
    import os
    wb = Workbook()
    ws = wb.active
    ws.title = "测试图片"
    # 保存测试文件
    test_output_path = "test_image_excel.xlsx"
    wb.save(test_output_path)
    print(f"测试Excel文件已保存到: {test_output_path}")
    