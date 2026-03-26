import json
import os
import tempfile

import openpyxl
import pandas as pd
import pymysql


EXCEL_PATH = r"f:\0graduationproject\sgs_toy_mvp-bd\sgs_toy_mvp-feature-agent_bd\docs\玩具标签审核梳理forMVP1028_2.xlsx"


def normalize_excel(path: str) -> pd.DataFrame:
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    merged_ranges = list(ws.merged_cells.ranges)

    for merged_range in merged_ranges:
        ws.unmerge_cells(str(merged_range))

    for merged_range in merged_ranges:
        min_row, min_col, max_row, max_col = (
            merged_range.min_row,
            merged_range.min_col,
            merged_range.max_row,
            merged_range.max_col,
        )
        value = ws.cell(row=min_row, column=min_col).value
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                ws.cell(row=row, column=col).value = value

    temp_file = tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False)
    temp_path = temp_file.name
    temp_file.close()
    try:
        wb.save(temp_path)
        return pd.read_excel(temp_path)
    finally:
        if os.path.exists(temp_path):
            os.remove(temp_path)


def to_json_value(text: str) -> str:
    if text == "":
        return "null"
    return json.dumps(text, ensure_ascii=False)


def clamp(value: str, max_len: int) -> str:
    if len(value) <= max_len:
        return value
    return value[:max_len]


def main() -> None:
    df = normalize_excel(EXCEL_PATH)

    conn = pymysql.connect(
        host=os.getenv("MYSQL_HOST", "127.0.0.1"),
        port=int(os.getenv("MYSQL_PORT", "3306")),
        user=os.getenv("MYSQL_USER", "root"),
        password=os.getenv("MYSQL_PASSWORD", "123456"),
        database=os.getenv("MYSQL_DB", "intelligent_detection"),
        charset="utf8mb4",
        cursorclass=pymysql.cursors.DictCursor,
        autocommit=True,
    )

    inserted = 0
    updated = 0
    skipped = 0
    errors = 0

    try:
        with conn.cursor() as cur:
            for _, row in df.iterrows():
                chapter = str(row.iloc[0]) if pd.notna(row.iloc[0]) else ""
                title = str(row.iloc[1]) if pd.notna(row.iloc[1]) else ""
                check_method = str(row.iloc[2]) if pd.notna(row.iloc[2]) else ""
                requirement = str(row.iloc[3]) if pd.notna(row.iloc[3]) else ""
                precondition = str(row.iloc[4]) if pd.notna(row.iloc[4]) else ""
                age_range_label = str(row.iloc[5]) if pd.notna(row.iloc[5]) else ""
                review_content = str(row.iloc[6]) if pd.notna(row.iloc[6]) else ""
                llm_prompt = str(row.iloc[8]) if pd.notna(row.iloc[8]) else ""

                chapter = clamp(chapter, 50)
                title = clamp(title, 200)
                check_method = clamp(check_method, 100)
                age_range_label = clamp(age_range_label, 50)

                if not chapter and not title and not check_method:
                    skipped += 1
                    continue

                precondition_json = to_json_value(precondition)

                try:
                    cur.execute(
                        """
                        SELECT id FROM rule
                        WHERE chapter = %s AND title = %s AND check_method = %s
                        LIMIT 1
                        """,
                        (chapter, title, check_method),
                    )
                    existed = cur.fetchone()

                    if existed and existed.get("id"):
                        cur.execute(
                            """
                            UPDATE rule
                            SET requirement = %s,
                                precondition = CAST(%s AS JSON),
                                age_range_label = %s,
                                review_content = %s,
                                llm_prompt = %s
                            WHERE id = %s
                            """,
                            (
                                requirement,
                                precondition_json,
                                age_range_label,
                                review_content,
                                llm_prompt,
                                int(existed["id"]),
                            ),
                        )
                        updated += 1
                    else:
                        cur.execute(
                            """
                            INSERT INTO rule
                            (chapter, title, check_method, requirement, precondition,
                             age_range_label, review_content, llm_prompt)
                            VALUES (%s, %s, %s, %s, CAST(%s AS JSON), %s, %s, %s)
                            """,
                            (
                                chapter,
                                title,
                                check_method,
                                requirement,
                                precondition_json,
                                age_range_label,
                                review_content,
                                llm_prompt,
                            ),
                        )
                        inserted += 1
                except Exception as e:
                    errors += 1
                    print(f"ROW_ERROR: {e}")

        print(
            f"IMPORT_DONE inserted={inserted} updated={updated} "
            f"skipped={skipped} errors={errors}"
        )
    finally:
        conn.close()


if __name__ == "__main__":
    main()
